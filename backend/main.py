from collections import Counter
from datetime import datetime, timedelta, timezone
import hashlib
from typing import Annotated
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from urllib import parse
import secrets
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
import requests
from pydantic import BaseModel
import jwt
import json
import os
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from email_validator import validate_email
import networkx as nx
from .ms_form_calculate import calculate_ranking_result
from .config import settings

bearer_scheme = HTTPBearer()

app = FastAPI()

origins = [
    settings.frontend_url,
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=False,
    allow_methods=['*'],
    allow_headers=['*'],
)

class GoogleAuthCallback(BaseModel):
    code: str

class TokenData(BaseModel):
    sub: str
    name: str
    picture: str

class Column(BaseModel):
    name: str
    index: int

class Columns(BaseModel):
    ranking: list[Column]
    choice_single_answer: list[Column]
    default: list[Column]

class CalculateResultsRequest(BaseModel):
    user_list_hash: str | None = None
    voting_form_hash: str
    check_user_list: bool = False
    columns: Columns

class UserListDetails(BaseModel):
    filename: str | None
    file_sha256: str
    num_users: int
    uploaded_at: str
    uploaded_by: str

class VotingFormDetails(BaseModel):
    filename: str | None
    file_sha256: str
    columns: Columns
    num_responses: int
    uploaded_at: str
    uploaded_by: str

class Row(BaseModel):
    row_number: int
    row: tuple[str | int | float | datetime | None, ...]

def create_access_token(data: dict, expires_delta: timedelta | None = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=settings.access_token_expire_minutes)
    to_encode.update({'exp': expire})
    encoded_jwt = jwt.encode(to_encode, settings.access_token_secret, algorithm='HS256')
    return encoded_jwt

def get_spreadsheet_worksheet(file_path: str):
    wb = load_workbook(file_path)
    ws = wb.active or wb.worksheets[0]
    return ws

def get_spreadsheet_num_rows(ws: Worksheet):
    return ws.max_row

def guess_is_ranking_column(ws: Worksheet, col_i: int):
    try:
        candidate_sets = Counter()
        for row_i in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_i, column=col_i)
            value = cell.value
            if not value:
                continue
            
            if type(value) != str:
                return False
            if value[-1] != ';':
                return False
            
            candidates = value[:-1].split(';')
            candidates.sort()
            candidates = tuple(candidates)
            candidate_sets[candidates] += 1
        most_common_candidate_set = candidate_sets.most_common(1)[0]
        return len(most_common_candidate_set) > 1
    except:
        return False

def get_column_types(ws: Worksheet):
    MS_FORM_COLUMNS = ['ID', 'Start time', 'Completion time', 'Email', 'Name', 'Last modified time']
    columns = {
        'default': [],
        'ranking': [],
        'choice_single_answer': [],
        # 'choice_multiple_answer': [],
    }
    for col_i in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col_i).value
        if col_name in MS_FORM_COLUMNS:
            columns['default'].append({'name': col_name, 'index': col_i})
        elif guess_is_ranking_column(ws, col_i):
            columns['ranking'].append({'name': col_name, 'index': col_i})
        else:
            columns['choice_single_answer'].append({'name': col_name, 'index': col_i})
    return columns

def check_user_email_in_list(email: str, user_list: list[str]):
    try:
        emailinfo = validate_email(email, check_deliverability=False)
        if settings.user_email_domains and emailinfo.domain not in [d.lower() for d in settings.user_email_domains]:
            return False
        if emailinfo.local_part.lower() not in user_list:
            return False
        return True
    except:
        return email.lower() in user_list

@app.get("/api/auth/google")
def google_auth():
    # https://developers.google.com/identity/protocols/oauth2/web-server#httprest
    client_id = settings.google_client_id
    state = secrets.token_urlsafe(16)
    parameters = {
        'client_id': client_id,
        'redirect_uri': f'{settings.frontend_url}/auth/google',
        'response_type': 'code',
        'scope': 'openid profile email',
        'state': state,
    }
    return {
        'url': f'https://accounts.google.com/o/oauth2/v2/auth?{parse.urlencode(parameters)}',
        'state': state
    }

@app.post("/api/auth/google")
def google_auth_callback(data: GoogleAuthCallback):
    # https://developers.google.com/identity/openid-connect/openid-connect#exchangecode
    request = requests.post('https://oauth2.googleapis.com/token', data={
        'code': data.code,
        'client_id': settings.google_client_id,
        'client_secret': settings.google_client_secret,
        'redirect_uri': f'{settings.frontend_url}/auth/google',
        'grant_type': 'authorization_code',
    })
    if request.status_code != 200:
        raise HTTPException(status_code=request.status_code, detail='Failed to exchange for token from Google')

    id_token = request.json().get('id_token')
    claims = jwt.decode(id_token, options={'verify_signature': False}) # do not need to verify signature as it is directly from Google

    email = claims.get('email') # email may not be unique to a user and may be changed but it is fine for this use case
    email_verified = claims.get('email_verified')
    hd = claims.get('hd')
    if not email or not email_verified:
        raise HTTPException(status_code=401, detail='Email not found or not verified')
    if settings.authorized_emails or settings.authorized_hds:
        if email not in settings.authorized_emails and hd not in settings.authorized_hds:
            raise HTTPException(status_code=401, detail='User not authorized')
    
    token_data = TokenData(sub=email, name=claims.get('name'), picture=claims.get('picture'))
    access_token = create_access_token(token_data.model_dump())
    return {'access_token': access_token}

class User(BaseModel):
    sub: str
    name: str
    picture: str

def get_current_user(token: Annotated[HTTPAuthorizationCredentials, Depends(bearer_scheme)]):
    try:
        payload = jwt.decode(token.credentials, settings.access_token_secret, algorithms=['HS256'])
        token_data = TokenData.model_validate(payload)
    except:
        raise HTTPException(status_code=401, detail='Could not validate credentials')
    return User(**token_data.model_dump())

@app.get('/api/auth/profile')
def profile(current_user: Annotated[User, Depends(get_current_user)]):
    return current_user

@app.post('/api/admin/user-list')
def upload_user_list(
    file: UploadFile,
    current_user: Annotated[User, Depends(get_current_user)],
):
    with open('data/user_list.txt', 'wb') as f:
        f.write(file.file.read())
    try:
        with open('data/user_list.txt', 'r') as f:
            num_lines = sum(1 for line in f if line.strip())
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail='File does not seem to be a text file or contains non-Unicode characters')
    with open('data/user_list.txt', 'rb') as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()
    details = UserListDetails(
        filename=file.filename,
        num_users=num_lines, # does not check duplicates
        file_sha256=file_hash,
        uploaded_at=datetime.now(timezone.utc).isoformat(),
        uploaded_by=current_user.sub,
    )
    with open('data/user_list_details.json', 'w') as f:
        f.write(details.model_dump_json(indent=2) + '\n')
    return {'message': 'File uploaded'}

@app.get('/api/admin/user-list')
def get_user_list_details():
    if not os.path.exists('data/user_list_details.json'):
        raise HTTPException(status_code=404, detail='User list not found')
    with open('data/user_list_details.json', 'r') as f:
        details = UserListDetails.model_validate(json.load(f))
    return details

@app.delete('/api/admin/user-list')
def delete_user_list():
    if os.path.exists('data/user_list_details.json'):
        os.remove('data/user_list_details.json')
    if os.path.exists('data/user_list.txt'):
        os.remove('data/user_list.txt')
    return {'message': 'User list deleted'}

@app.post('/api/admin/voting-form')
def upload_voting_form(
    file: UploadFile,
    current_user: Annotated[User, Depends(get_current_user)],
):
    with open('data/voting_form.xlsx', 'wb') as f:
        f.write(file.file.read())
    with open('data/voting_form.xlsx', 'rb') as f:
        file_hash = hashlib.sha256(f.read()).hexdigest()
    # get spreadsheet info
    try:
        ws = get_spreadsheet_worksheet('data/voting_form.xlsx')
    except:
        raise HTTPException(status_code=400, detail='Error occurred, maybe file is not a valid .xlsx file')
    columns = get_column_types(ws)
    num_responses = get_spreadsheet_num_rows(ws) - 1
    details = VotingFormDetails(
        filename=file.filename,
        file_sha256=file_hash,
        columns=Columns(**columns),
        num_responses=num_responses,
        uploaded_at=datetime.now(timezone.utc).isoformat(),
        uploaded_by=current_user.sub,
    )
    with open('data/voting_form_details.json', 'w') as f:
        f.write(details.model_dump_json(indent=2) + '\n')
    return {'message': 'File uploaded'}

@app.get('/api/admin/voting-form')
def get_voting_form_details():
    if not os.path.exists('data/voting_form_details.json'):
        raise HTTPException(status_code=404, detail='Voting form not found')
    with open('data/voting_form_details.json', 'r') as f:
        details = VotingFormDetails.model_validate(json.load(f))
    return details

@app.delete('/api/admin/voting-form')
def delete_voting_form():
    if os.path.exists('data/voting_form_details.json'):
        os.remove('data/voting_form_details.json')
    if os.path.exists('data/voting_form.xlsx'):
        os.remove('data/voting_form.xlsx')
    return {'message': 'Voting form deleted'}

@app.post('/api/admin/calculate-results')
def calculate_results(data: CalculateResultsRequest):
    warnings = []

    # check voting response exists
    if not os.path.exists('data/voting_form.xlsx') or not os.path.exists('data/voting_form_details.json'):
        raise HTTPException(status_code=404, detail='Voting form not found')
    
    # load voting response and check hash
    try:
        voting_form = get_spreadsheet_worksheet('data/voting_form.xlsx')
    except:
        raise HTTPException(status_code=400, detail='Error occurred, could not open voting response file')
    with open('data/voting_form_details.json', 'r') as f:
        voting_form_details = VotingFormDetails.model_validate(json.load(f))
    if voting_form_details.file_sha256 != data.voting_form_hash:
        warnings.append('Voting form hash does not match')

    # load user list if needed
    user_list = None
    if data.check_user_list:
        if not os.path.exists('data/user_list.txt') or not os.path.exists('data/user_list_details.json'):
            warnings.append('User list not found, skipping user list check')
        else:
            with open('data/user_list.txt', 'r') as f:
                user_list = [line.strip().lower() for line in f if line.strip()]
            with open('data/user_list_details.json', 'r') as f:
                user_list_details = UserListDetails.model_validate(json.load(f))
            if user_list_details.file_sha256 != data.user_list_hash:
                warnings.append('User list hash does not match')

    responses = [Row(row_number=i, row=row) for i, row in enumerate(voting_form.iter_rows(values_only=True, min_row=2), start=2)]
    if user_list:
        email_col_i = None
        for col_i in range(1, voting_form.max_column + 1):
            col_name = voting_form.cell(row=1, column=col_i).value
            if col_name == 'Email':
                email_col_i = col_i
                break
        if email_col_i is None:
            warnings.append('Email column not found in voting form, skipping user list check')
        else:
            responses = [row for row in responses if isinstance(row.row[email_col_i - 1], str) and check_user_email_in_list(str(row.row[email_col_i - 1]), user_list)]

    ranking_column_indices = [col.index for col in data.columns.ranking]
    choice_single_answer_column_indices = [col.index for col in data.columns.choice_single_answer]

    ranking_column_results = []
    for col_i in ranking_column_indices:
        errors = []
        warnings = []
        num_votes = 0
        num_abstain = 0
        num_invalid = 0

        column_responses = [(row.row_number, row.row[col_i - 1]) for row in responses]
        # remove non-strings
        column_responses_ = []
        for row_number, value in column_responses:
            if value is None or isinstance(value, str):
                column_responses_.append((row_number, value))
            else:
                num_invalid += 1
                warnings.append(f'Row {row_number} is not string, invalid and ignored')

        result_, warnings_, errors_ = calculate_ranking_result(column_responses_)
        warnings += warnings_
        errors += errors_

        if not result_:
            num_invalid = 0
            winners = None
            pairs = None
            lock_graph = None
        else:
            winners, pairs, lock_graph, num_votes_, num_abstain_, num_invalid_ = result_
            num_votes += num_votes_
            num_abstain += num_abstain_
            num_invalid += num_invalid_
            pairs = [pair.model_dump() for pair in pairs]
            lock_graph = nx.node_link_data(lock_graph, edges='edges') # type: ignore

        result = {
            'column_name': voting_form.cell(row=1, column=col_i).value,
            'winners': winners,
            'pairs': pairs,
            'lock_graph': lock_graph,
            'num_votes': num_votes,
            'num_abstain': num_abstain,
            'num_invalid': num_invalid,
            'errors': errors,
            'warnings': warnings,
        }
        ranking_column_results.append(result)

    choice_column_results = []
    for col_i in choice_single_answer_column_indices:
        column_responses = [(row.row_number, row.row[col_i - 1]) for row in responses]
        # print(column_responses)
    
    results = {
            'voting_form': {
                'filename': voting_form_details.filename,
                'file_sha256': voting_form_details.file_sha256,
                'uploaded_at': voting_form_details.uploaded_at,
                'uploaded_by': voting_form_details.uploaded_by,
            },
            'user_list': {
                'filename': user_list_details.filename,
                'file_sha256': user_list_details.file_sha256,
                'uploaded_at': user_list_details.uploaded_at,
                'uploaded_by': user_list_details.uploaded_by,
            } if user_list else None,
            'num_responses': voting_form_details.num_responses,
            'num_valid_responses': len(responses),
            'rank_column_results': ranking_column_results,
        }

    with open('data/results.json', 'w') as f:
        f.write(json.dumps(results, indent=2) + '\n')

    return {
        'results': results,
        'warnings': warnings,
    }
